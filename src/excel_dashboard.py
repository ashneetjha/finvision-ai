import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference

file_path = "data/output/payments.xlsx"

wb = load_workbook(file_path)
ws = wb.active

chart = BarChart()
chart.title = "Payment Status"
chart.y_axis.title = "Payable (1 = Yes, 0 = No)"
chart.x_axis.title = "File"

data = Reference(ws, min_col=5, min_row=1, max_row=ws.max_row)
cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)

chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)

ws.add_chart(chart, "H2")
wb.save(file_path)

print("Excel dashboard updated.")
