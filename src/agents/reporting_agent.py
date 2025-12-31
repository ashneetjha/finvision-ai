import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

class ReportingAgent:
    def __init__(self, output_dir="data/output"):
        self.output_dir = output_dir
        os.makedirs(self.output_dir, exist_ok=True)

    def generate_dashboard(self, df, stats):
        """
        Generates two files:
        1. ocr_data.xlsx (Raw OCR Output)
        2. FinVision_Dashboard.xlsx (Executive Dashboard with Colors)
        """
        dashboard_path = os.path.join(self.output_dir, "FinVision_Dashboard.xlsx")
        ocr_path = os.path.join(self.output_dir, "ocr_data.xlsx")
        
        # --- 1. SAVE RAW OCR DATA (SEPARATE FILE) ---
        if not df.empty:
            df.to_excel(ocr_path, index=False)
            print(f" [Reporting Agent] Raw OCR Data saved: {ocr_path}")
        
        # --- 2. CREATE DASHBOARD WORKBOOK ---
        with pd.ExcelWriter(dashboard_path, engine='openpyxl') as writer:
            # Create Sheets
            pd.DataFrame().to_excel(writer, sheet_name="Executive Summary")
            
            if not df.empty:
                df.to_excel(writer, sheet_name="Audit Logs", index=False)
            else:
                pd.DataFrame(["No Data Found"]).to_excel(writer, sheet_name="Audit Logs", header=False)

        # --- 3. APPLY STYLING TO DASHBOARD ---
        wb = load_workbook(dashboard_path)
        
        # Build Summary Sheet
        self._build_executive_summary(wb["Executive Summary"], stats)
        
        # Format Data Sheet (Red/Green logic)
        if not df.empty and "Audit Logs" in wb.sheetnames:
            self._format_audit_logs(wb["Audit Logs"], df)

        wb.save(dashboard_path)
        print(f" [Reporting Agent] Dashboard generated: {dashboard_path}")
        return dashboard_path

    def _build_executive_summary(self, ws, stats):
        # Title
        ws["A1"] = "FINVISION AI – EXECUTIVE DASHBOARD"
        ws["A1"].font = Font(size=18, bold=True, color="1F4E78")
        ws.merge_cells("A1:E1")

        # Metrics
        total = stats.get("total_rows", 0)
        risk = stats.get("unsigned_count", 0)
        safe = stats.get("verified_count", 0)
        ink = stats.get("ink_density", 0)
        
        metrics = [
            ("Total Logs Scanned", total, "E0E0E0"),
            ("Verified / Safe", safe, "C6EFCE"),
            ("Risk / Unsigned", risk, "FFC7CE"),
            ("Ink Density Detected", f"{ink}", "FFEB9C")
        ]

        row_start = 3
        for i, (title, val, color) in enumerate(metrics):
            # Label
            ws.cell(row=row_start + i, column=1, value=title).font = Font(bold=True, size=12)
            # Value
            cell_val = ws.cell(row=row_start + i, column=2, value=val)
            cell_val.font = Font(bold=True, size=12)
            cell_val.alignment = Alignment(horizontal="center")
            cell_val.fill = PatternFill("solid", fgColor=color)
            cell_val.border = Border(top=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'), bottom=Side(style='thin'))

        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 15

    def _format_audit_logs(self, ws, df):
        # Styles
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        red_font = Font(color="9C0006")
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        green_font = Font(color="006100")
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        # Header
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Logic
        try:
            status_col_idx = df.columns.get_loc("Audit Status") + 1
            for row in ws.iter_rows(min_row=2):
                status_val = row[status_col_idx - 1].value
                if status_val == "Risk (Unsigned/Empty)":
                    for cell in row:
                        cell.fill = red_fill
                        cell.font = red_font
                else:
                    for cell in row:
                        cell.fill = green_fill
                        cell.font = green_font
        except KeyError:
            pass

        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 18